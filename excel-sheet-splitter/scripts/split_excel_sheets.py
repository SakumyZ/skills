#!/usr/bin/env python3
"""
Excel Sheet Splitter - Split an Excel file into separate files per sheet
"""
import argparse
import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def split_excel_sheets(input_file: str, output_dir: str = None, prefix: str = None) -> dict:
    """
    Split an Excel file into separate files, one per sheet.

    Args:
        input_file: Path to the input Excel file
        output_dir: Output directory (default: same as input file)
        prefix: Prefix for output filenames (default: input filename)

    Returns:
        dict with 'status', 'files', and 'errors' keys
    """
    try:
        input_path = Path(input_file)

        if not input_path.exists():
            return {
                'status': 'error',
                'message': f'Input file not found: {input_file}',
                'files': [],
                'errors': [f'File not found: {input_file}']
            }

        # Set default output directory
        if output_dir is None:
            output_dir = input_path.parent
        else:
            output_dir = Path(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)

        # Set default prefix
        if prefix is None:
            prefix = input_path.stem

        # Load the workbook
        print(f"Loading workbook: {input_file}")
        wb = load_workbook(input_file)

        if len(wb.sheetnames) == 0:
            return {
                'status': 'error',
                'message': 'No sheets found in the workbook',
                'files': [],
                'errors': ['Workbook contains no sheets']
            }

        created_files = []
        errors = []

        # Process each sheet
        for sheet_name in wb.sheetnames:
            try:
                print(f"Processing sheet: {sheet_name}")

                # Create a new workbook for this sheet
                new_wb = Workbook()
                new_wb.remove(new_wb.active)  # Remove default sheet

                # Copy the source sheet
                source_sheet = wb[sheet_name]
                target_sheet = new_wb.create_sheet(sheet_name)

                # Copy all cells with values, formulas, and formatting
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet[cell.coordinate]

                        # Copy value or formula
                        if cell.value is not None:
                            if cell.data_type == 'f':  # Formula
                                target_cell.value = cell.value
                            else:
                                target_cell.value = cell.value

                        # Copy formatting
                        if cell.has_style:
                            target_cell.font = cell.font.copy()
                            target_cell.border = cell.border.copy()
                            target_cell.fill = cell.fill.copy()
                            target_cell.number_format = cell.number_format
                            target_cell.protection = cell.protection.copy()
                            target_cell.alignment = cell.alignment.copy()

                # Copy column widths
                for col in source_sheet.column_dimensions:
                    if source_sheet.column_dimensions[col].width:
                        target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width

                # Copy row heights
                for row in source_sheet.row_dimensions:
                    if source_sheet.row_dimensions[row].height:
                        target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

                # Copy merged cells
                for merged_range in source_sheet.merged_cells.ranges:
                    target_sheet.merge_cells(str(merged_range))

                # Generate output filename (sanitize sheet name for filesystem)
                safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
                output_filename = f"{prefix}_{safe_sheet_name}.xlsx"
                output_path = output_dir / output_filename

                # Save the new workbook
                new_wb.save(output_path)
                created_files.append(str(output_path))
                print(f"Created: {output_path}")

            except Exception as e:
                error_msg = f"Error processing sheet '{sheet_name}': {str(e)}"
                errors.append(error_msg)
                print(f"ERROR: {error_msg}", file=sys.stderr)

        # Prepare result
        if errors:
            status = 'partial_success' if created_files else 'error'
            message = f"Completed with {len(errors)} error(s). Created {len(created_files)} file(s)."
        else:
            status = 'success'
            message = f"Successfully split {len(created_files)} sheet(s)."

        return {
            'status': status,
            'message': message,
            'input_file': str(input_path),
            'output_directory': str(output_dir),
            'files': created_files,
            'errors': errors,
            'total_sheets': len(wb.sheetnames),
            'successful': len(created_files),
            'failed': len(errors)
        }

    except Exception as e:
        return {
            'status': 'error',
            'message': f'Failed to process Excel file: {str(e)}',
            'files': [],
            'errors': [str(e)]
        }


def main():
    parser = argparse.ArgumentParser(
        description='Split an Excel file into separate files, one per sheet'
    )
    parser.add_argument(
        'input_file',
        help='Path to the input Excel file'
    )
    parser.add_argument(
        '-o', '--output-dir',
        help='Output directory (default: same directory as input file)'
    )
    parser.add_argument(
        '-p', '--prefix',
        help='Prefix for output filenames (default: input filename without extension)'
    )
    parser.add_argument(
        '--json',
        action='store_true',
        help='Output result in JSON format'
    )

    args = parser.parse_args()

    # Run the split operation
    result = split_excel_sheets(args.input_file, args.output_dir, args.prefix)

    # Output results
    if args.json:
        import json
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print(f"\n{result['message']}")
        if result['files']:
            print("\nCreated files:")
            for file in result['files']:
                print(f"  - {file}")
        if result['errors']:
            print("\nErrors:")
            for error in result['errors']:
                print(f"  - {error}")

    # Exit with appropriate code
    sys.exit(0 if result['status'] in ['success', 'partial_success'] else 1)


if __name__ == '__main__':
    main()
