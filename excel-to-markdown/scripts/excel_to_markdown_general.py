#!/usr/bin/env python3
"""
excel_to_md.py

Excel (.xlsx) -> Markdown / HTML
支持：
- 合并单元格 fill / tl
- 日志
- 真实列优化
- 删除线内容自动忽略
"""

import argparse
import openpyxl
from openpyxl.utils import range_boundaries
import os
import re
import logging
import time
from html import escape


# ========= logging =========

logger = logging.getLogger("excel_to_md")

def setup_logging(verbose: bool):
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S"
    )


# ========= cell formatting =========

def get_cell_display(cell):
    """
    转换单元格为字符串
    - 删除线内容直接忽略
    - 支持粗体 / 斜体 / 超链接
    - 换行转 <br>
    """

    # ❶ 如果是删除线，直接忽略
    try:
        f = cell.font
        if f and getattr(f, "strike", False):
            return ""
    except Exception:
        pass

    val = cell.value
    if val is None:
        text = ""
    else:
        text = str(val).replace("\r", "").replace("\n", "<br>")

    # 超链接
    try:
        hl = cell.hyperlink
        if hl:
            url = hl.target if hasattr(hl, "target") and hl.target else str(hl)
            display_text = text if text else url
            text = f"[{display_text}]({url})"
    except Exception:
        pass

    # 避免破坏 markdown 表格
    text = text.replace("|", "\\|")

    # 字体样式
    try:
        f = cell.font
        if f:
            if getattr(f, "bold", False):
                text = f"**{text}**"
            if getattr(f, "italic", False):
                text = f"_{text}_"
    except Exception:
        pass

    return text


# ========= merged cell handling =========

def analyze_merged_cells(ws, merge_mode="tl"):
    placement = {}

    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))
        top_cell = ws.cell(min_row, min_col)
        top_text = get_cell_display(top_cell)

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if merge_mode == "fill":
                    placement[(r, c)] = top_text
                else:
                    if r == min_row and c == min_col:
                        placement[(r, c)] = top_text

    return placement


# ========= real max column optimization =========

def get_real_max_column(ws, scan_rows=200):
    max_col = 0
    max_row = ws.max_row or 0
    max_scan = min(max_row, scan_rows)

    for row in ws.iter_rows(min_row=1, max_row=max_scan):
        for cell in row:
            if cell.value not in (None, ""):
                if cell.column > max_col:
                    max_col = cell.column

    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col_m, max_row_m = range_boundaries(str(merged))
        if max_col_m > max_col:
            max_col = max_col_m

    return max_col


# ========= sheet -> rows =========

def sheet_to_table(ws, merge_mode="tl", trim=True):
    start_time = time.time()

    max_r = ws.max_row or 0
    real_max_c = get_real_max_column(ws)
    max_c = real_max_c

    logger.info(
        f"Sheet '{ws.title}' 开始处理 "
        f"(行={max_r}, 真实列={real_max_c})"
    )

    placement = analyze_merged_cells(ws, merge_mode=merge_mode)

    rows = []

    for r in range(1, max_r + 1):

        if max_r >= 1000 and r % 100 == 0:
            logger.info(f"Sheet '{ws.title}' 进度: {r}/{max_r}")

        row_list = []

        for c in range(1, max_c + 1):
            if (r, c) in placement:
                cell_text = placement[(r, c)]
            else:
                cell = ws.cell(r, c)
                cell_text = get_cell_display(cell)

            row_list.append(cell_text)

        rows.append(row_list)

    duration = time.time() - start_time
    logger.info(
        f"Sheet '{ws.title}' 处理完成，用时 {duration:.2f} 秒"
    )

    if trim:
        rows = trim_empty_rows_cols(rows)

    return rows


def trim_empty_rows_cols(rows):
    if not rows:
        return rows

    top = 0
    bottom = len(rows) - 1

    while top <= bottom and all(
        (not cell or str(cell).strip() == "")
        for cell in rows[top]
    ):
        top += 1

    while bottom >= top and all(
        (not cell or str(cell).strip() == "")
        for cell in rows[bottom]
    ):
        bottom -= 1

    if top > bottom:
        return []

    rows = rows[top:bottom + 1]

    num_cols = len(rows[0])
    last_col = -1

    for c in range(num_cols - 1, -1, -1):
        if any(
            rows[r][c] and str(rows[r][c]).strip() != ""
            for r in range(len(rows))
        ):
            last_col = c
            break

    if last_col == -1:
        return []

    rows = [row[:last_col + 1] for row in rows]
    return rows


# ========= rendering =========

def rows_to_markdown(rows):
    if not rows:
        return ""

    header = rows[0]
    sep = ["---"] * len(header)

    def fmt_row(r):
        cells = ["" if c is None else str(c) for c in r]
        return "| " + " | ".join(cells) + " |"

    lines = []
    lines.append(fmt_row(header))
    lines.append("| " + " | ".join(sep) + " |")

    for r in rows[1:]:
        lines.append(fmt_row(r))

    return "\n".join(lines)


# ========= workbook conversion =========

def convert_workbook(input_xlsx, output_path,
                     merge_mode="tl"):

    wb = openpyxl.load_workbook(input_xlsx, data_only=True)

    logger.info(f"开始处理工作簿: {input_xlsx}")
    logger.info(f"Sheet 总数: {len(wb.sheetnames)}")

    combined = []

    for idx, sheetname in enumerate(wb.sheetnames, start=1):
        logger.info(f"======== Sheet {idx}/{len(wb.sheetnames)} ========")

        ws = wb[sheetname]
        rows = sheet_to_table(ws, merge_mode=merge_mode, trim=True)

        if not rows:
            content = f"<!-- sheet {sheetname} is empty -->\n"
        else:
            md_table = rows_to_markdown(rows)
            content = f"## {sheetname}\n\n{md_table}\n"

        combined.append(content)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(combined))

    logger.info(f"全部完成，输出文件: {output_path}")


# ========= CLI =========

def main():
    p = argparse.ArgumentParser(description="Excel to Markdown")
    p.add_argument("input", help="输入 .xlsx 文件")
    p.add_argument("-o", "--output", help="输出文件路径")
    p.add_argument("--merge-mode",
                   choices=["fill", "tl"],
                   default="tl")
    p.add_argument("--verbose",
                   action="store_true")

    args = p.parse_args()

    setup_logging(args.verbose)

    if not os.path.exists(args.input):
        logger.error("输入文件不存在")
        return

    output = args.output
    if not output:
        base = os.path.splitext(os.path.basename(args.input))[0]
        output = base + ".md"

    convert_workbook(
        args.input,
        output,
        merge_mode=args.merge_mode
    )


if __name__ == "__main__":
    main()